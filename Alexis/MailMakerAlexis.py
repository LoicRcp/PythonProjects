import os
import openpyxl
import xlrd


wb = xlrd.open_workbook('Octobre 2020 - MODIFIER.xls')
sheet = wb.sheet_by_index(0)

print(sheet.nrows)

variables = ['sex','nom','ville','mail']

clientList = []
for i in range(1, sheet.nrows):
    if sheet.cell(i, 15) != 'OK':
        client = {}
        client[variables[0]] = sheet.cell(i,2).value
        client[variables[1]] = sheet.cell(i,4).value
        client[variables[2]] = sheet.cell(i,9).value
        client[variables[3]] = sheet.cell(i,11).value
        clientList.append(client)
        print(client)



